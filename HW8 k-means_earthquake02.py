# -*- coding: utf-8 -*-
"""
Created on Tue May  8 00:04:30 2018

@author: Ben
"""

import numpy as np# 引用 numpy 套件並縮寫為 np
import matplotlib.pyplot as plt# 引用 matplotlib.pyplot 套件並縮寫為 plt
from matplotlib import style# 引用 style 套件
style.use("ggplot")# 使用 ggplot
from sklearn.cluster import KMeans# 引用 KMeans 套件
from scipy.spatial.distance import cdist# 引用 cdist 套件
import pandas as pd# 引用 pandas 套件並縮寫為 pd

from openpyxl import load_workbook# 引用套件 load_workbook
wb = load_workbook("F14041088.xlsx")# 載入 excel 檔
sheet = wb.get_sheet_by_name("Sheet1")# 得到 sheet1 的資料

G = []
H = []
I = []
seismic = []
num = []

for i in sheet["G"]:# 讀取 G 行資料
    G.append(i.value)
for i in sheet["H"]:# 讀取 H 行資料
    H.append(i.value)
for i in sheet["I"]:# 讀取 I 行資料
    I.append(i.value)
    
G.remove('6.經度')# 去除不要的資料
H.remove('7.緯度')# 去除不要的資料
I.remove('8.規模')# 去除不要的資料

j = 0
for i in I:
    if float(i) >= 4:# 選出規模大於 4 的地震
        seismic.append(i)
        num.append(j)# 規模大於 4 的資料編號
    j+=1
X = np.zeros((len(G),2))
x = []
y = []
for i in num:    
    x.append(float(G[i]))
for i in num:
    y.append(float(H[i]))
X = np.stack((x,y),axis=-1)# 將一維陣列 x 和 y 組成一個二維陣列

colors = ['b.','g.','r.','c.','m.','y.','k.','w.','b.','g.']# 顏色  
ran = [3,4,5,6,7,8,9,10]# k 的範圍

Dis = np.zeros((max(ran),max(ran)))# 每群的鑑別率
Agg = []# 每群的凝聚率
Sum_Dis = 0# 每群鑑別率的總和
Sum_dk = 0# 1/Dk 的總和

j = 0# 計數
j1 = 0# 計數

mean1 = np.zeros((max(ran),max(ran)))# 每群資料到群中心的平均距離
sd1 = np.zeros((max(ran),max(ran)))# 每群資料到群中心的標準差
data = np.zeros((len(X),2,max(ran)))# 按照所屬群分類的資料

n_min = np.zeros((max(ran),max(ran)))# 兩群間最小距離
k_list = []# 紀錄 k 值
Eva = []# 總評估值
for k in ran:# 每次分成不同個數的群
    #歸零
    data = np.zeros((len(X),2,max(ran)))# 按照所屬群分類的資料
    r_cent = np.zeros((len(X),max(ran)))# 到所屬群的群中心距離
    q = [0] * max(ran)# 每一群所含的資料數
    Sum_Dis = 0# 每群鑑別率的總和
    #聚類
    kmeans = KMeans(n_clusters=k)# 分成 k 群
    kmeans.fit(X)# 分群
    centroids = kmeans.cluster_centers_# 計算群中心
    labels = kmeans.labels_# 得到分群標籤
     
    d_dist = cdist(X, centroids, 'euclidean')# 計算所有資料到所有群中心的距離
    c_dist = cdist(centroids, centroids, 'euclidean')# 計算所有群中心相互的距離
        
    for i in range(k):
        j = 0
        for lab in labels:#按照標籤分類資料          
            if lab == i:
                data[q[i],0,i] = x[j]# 按照所屬群分類的資料的 x 值
                data[q[i],1,i] = y[j]# 按照所屬群分類的資料的 y 值
                r_cent[q[i],i] = d_dist[j,i]# 到所屬群的群中心距離
                
                q[i] = q[i] + 1 # 計算每一群所含的資料數
            j+=1 
         
    for iy in range(k):# 每一群
        mean1[iy,k-1] =np.mean(r_cent[:(q[iy]),iy]) # 每群資料到群中心的平均距離
        sd1[iy,k-1] =np.std(r_cent[:(q[iy]),iy]) # 每群資料到群中心的標準差
        
        t1 = 0
        t2 = 0
        j = 0 
        for i2 in d_dist[:,iy]:
            if labels[j] == iy and i2 <= mean1[iy,k-1]:
                t1+=1    # 計算在平均內的資料數
            if labels[j] == iy and i2 <= mean1[iy,k-1] + 3 * sd1[iy,k-1]:
                t2+=1   # 計算在 3 個標準差內的資料數
            j+=1
        if t2 != 0:
            Agg.append(t1/t2) # 存入凝聚率
            
        if np.max(r_cent[:,iy]) != 0:
            dk = q[iy]/((np.max(r_cent[:,iy]))**2)# 計算每一群的密度
        Sum_dk = Sum_dk + dk# dk 的總和
        
        for ix in range(k):
            n_min[iy,ix] = (np.min(cdist(data[:q[iy],:,iy],data[:q[ix],:,ix], 'euclidean')))# 計算兩群間最小距離
            if c_dist[iy,ix] != 0:
                Dis[iy,ix] = n_min[iy,ix]/c_dist[iy,ix]# 計算鑑別率
                
        for i in range(len(X)): 
            plt.plot(X[i][0], X[i][1], colors[labels[i]], markersize = 10)# 用不同的顏色畫上每個已分群的點

    Avg_Agg = (np.mean(Agg[:]))# 計算分成 k 群時的平均凝聚率
    Avg_Dis = ((sum(sum(Dis)))/(k*(k-1)))# 計算分成 k 群時的平均群鑑別率
    Avg_dk = Sum_dk/k # 計算分成 k 群時的平均 dk
    Eva.append(0.6*Avg_Dis + 0.8*Avg_Agg + 0.0005*Avg_dk)# 計算總評估值
    k_list.append(k)# 紀錄 k 值
    plt.scatter(centroids[:, 0],centroids[:,1], marker = "x", s=150, linewidths = 5,zorder = 10)#用 X 標示套件產生的群中心
    plt.show()# 顯示圖表
df=pd.DataFrame({'k':k_list,'總評估值':Eva})# 畫成表格
print(df)# 顯示表格